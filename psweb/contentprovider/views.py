from django.views.generic import TemplateView, FormView
from django.forms.models import formset_factory
from http.client import HTTPSConnection
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render, redirect
from .models import *
from .forms import OfflineContentloadForm, BulkCourseForm
from learn.models import Course, CourseTag, CourseProvider, DifficultyChoice, Instructor, Duration, LiveTraining
from django.contrib.auth.mixins import LoginRequiredMixin
from datetime import timedelta
from django.utils import timezone
from django.template import loader, Context
from base64 import b64encode
import ssl
import requests
import psycopg2
import json
import nltk
from nltk.corpus import stopwords
from functools import reduce
from openpyxl import load_workbook
import logging
import os
import time
from owlready2 import get_ontology, onto_path

onto_path.append(os.path.dirname(__file__))
onto = get_ontology("CSO.owl")

userAndPass = b64encode(b"GpxNVedkBslJE6CTga0f56iRG4vzzmYU24gzH0g5:FGMx5x8Vjr7LyBokikzIT9t4uFSSa30HMhMGcEHZBy38FV2snjwew0l9o3ctugs1KRcIvBQyZDidYKuMKrWUGHCA0qRNYMvFg859QhpatbpBPZW3QNAeJzpHBAYNkBoy").decode("ascii")
headers = { 'Authorization' : 'Basic %s' %  userAndPass }
nltk.download(info_or_id='stopwords', download_dir=os.path.dirname(__file__))

PLURALSIGHT_EXTRACTOR_VERSION = 1
UDEMY_EXTRACTOR_VERSION = 1
LYNDA_EXTRACTOR_VERSION = 1
COURSERA_EXTRACTOR_VERSION = 1
OFFLINE_EXTRACTOR_VERSION = 1

def AddCourseTags(course):
    tagsSource = course.title
    for tag in GetTags(tagsSource):
        defaults = {'priority': 100}
        categorytag, created = CourseTag.objects.get_or_create(name=tag,
                                                               defaults=defaults)
        if created:
            categorytag.save()
        course.tags.add(categorytag)


def GetTags(sourceStr):
    onto.load()
    stop_words = set(stopwords.words('english'))
    cleaned_words = ""
    repls = ('.', ''), ('&', '')
    sourceStr = reduce(lambda a, kv: a.replace(*kv), repls, sourceStr)
    for word in sourceStr.split():
        if word not in stop_words:
            cleaned_words += " " + word
    result_wordpairs = []
    previous_word = None
    p1tags = []
    for word in cleaned_words.split():
        if not word[:1].isalnum():
            previous_word = None
        else:
            word_lower = word.lower()
            if previous_word:
                result_wordpairs.append(previous_word + '%20' + word_lower)
            previous_word = word_lower
    for wordpair in result_wordpairs:
        if onto.search(iri = "*" + wordpair):
            newtag = wordpair.replace("%20", ' ')
            p1tags.append(newtag)
            cleaned_words.replace(newtag, '')
    for tag in cleaned_words.split():
        if onto.search(iri = "*" + tag):
            p1tags.append(tag)
    return p1tags


class UdemyImport(LoginRequiredMixin, TemplateView):
    def get(self, request, *args, **kwargs):
        courseProvider = CourseProvider.objects.get(name='Udemy')
        courses_processed = 0
        # for each category and each level, get all the courses
        for category in UdemyCourseCategory.objects.all():
            defaults = {'priority': 100}
            categorytag, created = CourseTag.objects.get_or_create(name=category.title, defaults = defaults)
            if created:
                categorytag.save()
            for level in [('beginner', DifficultyChoice.Beginner), ('intermediate', DifficultyChoice.Intermediate), ('expert', DifficultyChoice.Advanced), ('all', DifficultyChoice.All)]:
                for duration in Duration:
                    payload = {'category': category.title, 'instructional_level': level[0], 'page_size': 10000,
                               'duration': duration.value[0], 'language': 'en'}
                    page = 0
                    res = None
                    while True:
                        rawdata, created = UdemyRawData.objects.get_or_create(category=category, level=level[1].value,
                                                                              duration=duration.value[1], page=page)
                        if created or rawdata.raw_data == None or rawdata.raw_data == '':
                            # try 10 times in case of failure
                            skipToNextQuery = False
                            for ii in range (0, 10):
                                if res:
                                    if 'next' in res and res['next']:
                                        request = requests.request('GET', res['next'], headers=headers)
                                    else:
                                        skipToNextQuery = True
                                        break
                                else:
                                    request = requests.request('GET', 'https://www.udemy.com/api-2.0/courses', params=payload,
                                                           headers=headers)
                                new_res = request.json()
                                if 'results' in new_res:
                                    res = new_res
                                    rawdata.raw_data = json.dumps(new_res)
                                    rawdata.save()
                                    break
                                else:
                                    time.sleep(30)
                            if skipToNextQuery:
                                break

                        else:
                            res = json.loads(rawdata.raw_data)

                        try:
                            results = res['results']
                        except Exception:
                            pass

                        for courseitem in results:
                            title = courseitem['title'] if 'title' in courseitem else None
                            url = courseProvider.url + courseitem['url'] if 'url' in courseitem else None
                            try:
                                price = courseitem['price_detail']['amount']
                            except Exception as e:
                                price = 0

                            thumbnail = courseitem['image_480x270'] if 'image_480x270' in courseitem else None

                            if title and url:
                                defaults = {'difficulty': level[1].value}
                                course, created = Course.objects.get_or_create(course_id=courseitem['id'], defaults=defaults)
                                if not created:
                                    if UDEMY_EXTRACTOR_VERSION == course.extractor_version:
                                        continue
                                    course.tags.through.objects.all().delete()
                                    course.instructors.through.objects.all().delete()

                                course.title = title
                                course.url = url
                                course.provider = courseProvider
                                course.difficulty = level[1].value
                                course.status = 1
                                course.price = price
                                course.thumbnail = thumbnail
                                course.duration = timedelta(hours=duration.value[1])
                                course.save()
                                AddCourseTags(course)
                                course.tags.add(categorytag)
                                # add instructors
                                for instructor in courseitem['visible_instructors']:
                                    defaults = {'url' : courseProvider.url + instructor['url'], 'photo' : instructor['image_100x100']}
                                    newInstructor, created = Instructor.objects.get_or_create(name=instructor['name'], defaults = defaults)
                                    if created:
                                        newInstructor.save()
                                    course.instructors.add(newInstructor)
                                course.extractor_version = UDEMY_EXTRACTOR_VERSION
                                course.save()
                            courses_processed += 1
                            print("Udemy course Processed count: " + str(courses_processed))

                        rawdata.processed = timezone.now()
                        rawdata.save()
                        page += 1

        return HttpResponse("Success")



#This sets up the https connection
c = HTTPSConnection("www.udemy.com", context=ssl._create_unverified_context())
#we need to base 64 encode it 
#and then decode it to acsii as python 3 stores it as a byte string
#then connect

# conn = psycopg2.connect(database='purpleskillsdb', user='psroot', password='pswhatever1', host='127.0.0.1', port='')
# udemyPrefix = "https://www.udemy.com"
# #print (data)
# cur = conn.cursor()
#
#
# cur.execute("INSERT INTO learn_courseprovider (name, status, logo) SELECT %s, %s, %s WHERE NOT EXISTS (SELECT 1 FROM learn_courseprovider where name = \'Udemy\') ", ("Udemy","0","https://www.udemy.com/staticx/udemy/images/v6/mstile-144x144.png"))
#
# for item in results:
#     my_data = {field: item[field] for field in fields}
#     cur.execute("INSERT INTO contentprovider_udemycourse (id,title,url) VALUES (%s, %s, %s) ON CONFLICT(id) DO NOTHING ", (my_data['id'],my_data['title'],udemyPrefix+my_data['url']))
#     cur.execute("INSERT INTO learn_course (title,url, provider_id) SELECT  %s, %s, ID FROM learn_courseprovider where name = \'Udemy\' AND  NOT EXISTS (SELECT 1 FROM learn_course where title=%s)", (my_data['title'],udemyPrefix+my_data['url'], my_data['title']))
#
#
#
#
# # commit changes
# conn.commit()
# # Close the connection
# conn.close()
import csv
import urllib3

class PluralSightImport(LoginRequiredMixin, TemplateView):
    levels = {}
    levels[0] = {'fundamentals', 'beginner', 'introduction', 'meet', 'getting started', 'introduces', 'basics'}
    levels[1] = {'advanced'}
    ignore_words = {'team-foundation', 'team foundation'}

    def computeDifficulty(self, courseId, courseTitle, courseDesc):
        courseTitle = courseTitle.lower()
        courseDesc = courseDesc.lower()
        # first remove ignore words
        for word_to_ignore in self.ignore_words:
            courseId = courseId.replace(word_to_ignore, '')
            courseTitle = courseTitle.replace(word_to_ignore, '')
            courseDesc = courseDesc.replace(word_to_ignore, '')

        courseId = courseId.split('-')
        courseTitle = courseTitle.split()
        courseDesc = courseDesc.split()

        if len(self.levels[0].intersection(courseId)) > 0 or len(self.levels[0].intersection(courseTitle)) > 0:
            return DifficultyChoice.Beginner.value

        if len(self.levels[1].intersection(courseId)) > 0 or len(self.levels[1].intersection(courseTitle)) > 0:
            return DifficultyChoice.Advanced.value

        if len(self.levels[0].intersection(courseDesc)) > 0:
            return DifficultyChoice.Beginner.value

        if len(self.levels[1].intersection(courseDesc)) > 0:
            return DifficultyChoice.Advanced.value

        return DifficultyChoice.Intermediate.value

    def get(self, request, *args, **kwargs):
        courseProvider = CourseProvider.objects.get(name='PluralSight')
        url = 'http://api.pluralsight.com/api-v0.9/courses'
        http = urllib3.PoolManager()
        response = http.request('GET', url)
        cr = csv.reader(response.data.decode("utf-8").splitlines())
        omittedFirstRow = False
        courses_processed = 0
        for courseitem in cr:
            if omittedFirstRow == False:
                omittedFirstRow = True
            else:
                if courseitem[6] == 'no':
                    difficulty = self.computeDifficulty(courseitem[0], courseitem[1], courseitem[4])
                    defaults = {'difficulty': difficulty}
                    course, created = Course.objects.get_or_create(course_id=courseitem[0], defaults=defaults)
                    if not created:
                        if PLURALSIGHT_EXTRACTOR_VERSION == course.extractor_version:
                            continue
                        course.tags.through.objects.all().delete()
                        course.instructors.through.objects.all().delete()
                    course.title = courseitem[1]
                    course.url = courseProvider.url + '/' + courseitem[0]
                    course.provider = courseProvider
                    course.difficulty = difficulty
                    course.status = 1
                    course.description = courseitem[4]
                    course.duration = timedelta(seconds=int(courseitem[2]))
                    course.save()
                    AddCourseTags(course)
                    course.extractor_version = PLURALSIGHT_EXTRACTOR_VERSION
                    course.save()
                courses_processed += 1
                print("PluralSight course Processed count: " + str(courses_processed))
        return HttpResponse("Success")


# https://api.coursera.org/api/courses.v1?start=0&limit=3&includes=instructorIds,partnerIds,specializations,s12nlds,v1Details,v2Details,instructors.v1,courseDerivativesV2,difficultyLevelTag&fields=instructorIds,partnerIds,specializations,s12nlds,description,display,photoUrl,description,v1Details,v2Details,instructors.v1(title),workload,domainTypes,categories
from html.parser import HTMLParser
import re
class CourseraImport(LoginRequiredMixin, TemplateView):
    course_difficulty = None

    class CourseraHTMLParser(HTMLParser):
        grab_course_data = False
        course_duration = None

        def handle_starttag(self, tag, attrs):
            if tag == 'script':
                attr_type = [v for i, v in enumerate(attrs) if v[0] == 'type']
                if len(attr_type) > 0 and attr_type[0][1] == 'application/ld+json':
                    self.grab_course_data = True

        def handle_endtag(self, tag):
            self.grab_course_data = False

        def handle_data(self, data):
            if self.grab_course_data == True:
                self.calc_course_props(data)

        def calc_course_props(self, data):
            json_data = json.loads(data)
            for graph_data in json_data['@graph']:
                if graph_data['@type'] == "Course":
                    time_data_matches = re.match("P((\d+)D)?(T((\d+)H)?((\d+)M)?)?", graph_data['timeRequired'])
                    total_secs = 0
                    try:
                        if time_data_matches.group(2):
                            total_secs += int(time_data_matches.group(2)) * 3600 * 24
                        if time_data_matches.group(5):
                            total_secs += int(time_data_matches.group(5)) * 3600
                        if time_data_matches.group(7):
                            total_secs += int(time_data_matches.group(7)) * 60
                    except Exception as e:
                        if total_secs > 0:
                            pass
                        else:
                            raise
                    self.course_duration = timedelta(seconds= total_secs)
                    assert (total_secs > 0)
                    break


    def get(self, request, *args, **kwargs):
        courseProvider = CourseProvider.objects.get(name='Coursera')
        url = 'https://api.coursera.org/api/courses.v1?start=0&limit=10&includes=instructorIds,partnerIds,specializations,s12nlds,v1Details,v2Details,instructors.v1(title)&fields=instructorIds,partnerIds,specializations,s12nlds,description,display,photoUrl,description,v1Details,v2Details,instructors.v1(title)'
        http = urllib3.PoolManager()
        response = http.request('GET', url)
        courses_processed = 0
        while True:
            # first parse the instructors
            json_data = json.loads(response.data.decode("utf-8"))
            for instructor in json_data['linked']['instructors.v1']:
                newInstructor, created = Instructor.objects.get_or_create(name=instructor['fullName'])
                if created:
                    newInstructor.save()

            for courseitem in json_data['elements']:
                self.course_difficulty = DifficultyChoice.Intermediate.value
                try:
                    course_difficulty_str = re.search('\"level\":(\")?(\w+)',raw_data).group(2).lower()
                    if (course_difficulty_str == "beginner"):
                        self.course_difficulty = DifficultyChoice.Beginner.value
                    elif (course_difficulty_str == "intermediate"):
                        self.course_difficulty = DifficultyChoice.Intermediate.value
                    elif (course_difficulty_str == "advanced"):
                        self.course_difficulty = DifficultyChoice.Advanced.value
                except Exception as e:
                    pass
                defaults = {'difficulty': self.course_difficulty}
                course, created = Course.objects.get_or_create(course_id=courseitem['id'], defaults=defaults)
                if not created:
                    if COURSERA_EXTRACTOR_VERSION == course.extractor_version:
                        continue
                    course.tags.through.objects.all().delete()
                    course.instructors.through.objects.all().delete()
                course_url = courseProvider.url + '/' + courseitem['slug']
                course_data, created = CourseraRawData.objects.get_or_create(course_id=courseitem['id'])
                if created or course_data.raw_data in [None, '']:
                    # get course details
                    response = http.request('GET', course_url)
                    raw_data = response.data.decode("utf-8")
                    course_data.raw_data = raw_data
                    course_data.retrieved = timezone.now()
                    course_data.save()
                raw_data = course_data.raw_data

                parser = self.CourseraHTMLParser()
                parser.feed(raw_data)

                course.title = courseitem['name']
                course.url = course_url
                course.provider = courseProvider
                course.status = 1
                course.description = courseitem['description']
                course.difficulty = self.course_difficulty
                course.duration = parser.course_duration
                course.save()
                # add course tags
                AddCourseTags(course)
                course_data.processed = timezone.now()
                course_data.save()

                course.extractor_version = COURSERA_EXTRACTOR_VERSION
                course.save()
                courses_processed += 1
                print("Coursera course Processed count: " + str(courses_processed))

            if 'next' in json_data['paging'] and json_data['paging']['next'] != None:
                url = 'https://api.coursera.org/api/courses.v1?limit=5000&includes=instructorIds,partnerIds,specializations,s12nlds,v1Details,v2Details,instructors.v1(title)&fields=instructorIds,partnerIds,specializations,s12nlds,description,display,photoUrl,description,v1Details,v2Details,instructors.v1(title)' + '&start=' + json_data['paging']['next']
                http = urllib3.PoolManager()
                response = http.request('GET', url)
            else:
                break
        return HttpResponse("Success")



import zipfile
import io
class LyndaImport(LoginRequiredMixin, TemplateView):
    def computeCourseDifficulty(self, level_str):
        if (level_str == 'Beginner'):
            return DifficultyChoice.Beginner.value
        elif (level_str == 'Intermediate'):
            return DifficultyChoice.Intermediate.value
        elif (level_str == 'Intermediate'):
            return DifficultyChoice.Intermediate.value
        else:
            return DifficultyChoice.Intermediate.value

    def computeDuration(self, duration_str):
        time_strs = duration_str.split(":")
        return int(time_strs[0]) * 3600 + int(time_strs[1]) * 60 + int(time_strs[2])

    def get(self, request, *args, **kwargs):
        courseProvider = CourseProvider.objects.get(name='Lynda')
        lyndaData = LyndaRawData.objects.all()
        if lyndaData.count() > 0:
            lyndaData = lyndaData.first()
            raw_data = lyndaData.raw_data
        else:
            url = 'http://www.lynda.com/courselist/'
            http = urllib3.PoolManager()
            response = http.request('GET', url)
            raw_data = response.data
            lyndaData = LyndaRawData(raw_data=raw_data)
            lyndaData.save()

        with zipfile.ZipFile(io.BytesIO(raw_data)) as lyndazip:
            cr = csv.reader(io.TextIOWrapper(lyndazip.open('lynda.com Courses.csv'), encoding='utf-8').read().splitlines())
            omittedFirstRow = False
            courseProcessedCount = 0
            for courseitem in cr:
                if omittedFirstRow == False:
                    omittedFirstRow = True
                else:
                    if courseitem[16] == 'active':
                        difficulty = self.computeCourseDifficulty(courseitem[4])
                        defaults = {'difficulty': difficulty}
                        course, created = Course.objects.get_or_create(course_id=courseitem[0], defaults = defaults)
                        if not created:
                            if LYNDA_EXTRACTOR_VERSION == course.extractor_version:
                                continue
                            course.tags.through.objects.all().delete()
                            course.instructors.through.objects.all().delete()
                        course.title = courseitem[1]
                        course.url = courseitem[10]
                        course.provider = courseProvider
                        course.difficulty = difficulty
                        course.status = 1
                        course.description = courseitem[9]
                        course.thumbnail = courseitem[15]
                        course.duration = timedelta(seconds=self.computeDuration(courseitem[5]))
                        course.save()
                        for instructor in courseitem[2].split(','):
                            newInstructor, created = Instructor.objects.get_or_create(name=instructor)
                            if created:
                                newInstructor.save()
                            course.instructors.add(newInstructor)
                        AddCourseTags(course)
                        course.extractor_version = LYNDA_EXTRACTOR_VERSION
                        course.save()
                        courseProcessedCount += 1
                        print("Lynda course Processed count: " + str(courseProcessedCount) )

        lyndaData.processed = timezone.now()
        lyndaData.save()
        return HttpResponse("Success")

class OfflineImport(LoginRequiredMixin, TemplateView):
    template_name = "offline_import.html"
    success_url = '/learn'
    # form_class=OfflineContentloadForm
    prefix="offc"

    def computeDuration(self, duration_str):
        time_strs = duration_str.split(":")
        return int(time_strs[0]) * 3600 + int(time_strs[1]) * 60 + int(time_strs[2])

    def get(self, request, *args, **kwargs):
        self.object = None
        offlinecontentform = OfflineContentloadForm()
        CourseFormSet = formset_factory(BulkCourseForm, extra=1, validate_max=True)
        offc_formset = CourseFormSet(prefix=self.prefix)
        return self.render_to_response(self.get_context_data(uploadform=offlinecontentform, formset=offc_formset))

    def post(self, request, *args, **kwargs):
        context = self.get_context_data(**kwargs)
        self.object = None
        CourseFormSet = formset_factory(BulkCourseForm, extra=1, validate_max=True)
        offc_formset = CourseFormSet(request.POST, prefix=self.prefix)
        # save the uploaded course data
        author = self.request.user
        if offc_formset.is_valid():
            for offc_form in offc_formset.forms:
                if offc_form.has_changed():
                    if offc_form.is_valid():
                        title = offc_form.cleaned_data["title"]
                        url = offc_form.cleaned_data["url"]
                        description = offc_form.cleaned_data["description"]
                        difficulty = offc_form.cleaned_data["difficulty"]
                        duration = offc_form.cleaned_data["duration"]
                        instructor_name = offc_form.cleaned_data["instructor_name"]
                        instructor_company = offc_form.cleaned_data["instructor_company"]
                        instructor_web = offc_form.cleaned_data["instructor_web"]
                        price = offc_form.cleaned_data["price"]
                        max_students = offc_form.cleaned_data["max_students"]
                        min_students = offc_form.cleaned_data["min_students"]
                        session_count = offc_form.cleaned_data["session_count"]
                        prerequisites = offc_form.cleaned_data["prerequisites"]
                        course = Course()
                        course.title = title
                        course.url = url
                        course.difficulty = difficulty
                        course.status = 1
                        course.mode=2 #offline
                        course.description = description
                        course.price = price
                        course.duration = duration
                        course.save()
                        for instructor in instructor_name.split(','):
                            newInstructor, created = Instructor.objects.get_or_create(name=instructor)
                            if created:
                                newInstructor.save()
                            course.instructors.add(newInstructor)
                        newProvider, p_created = CourseProvider.objects.get_or_create(name=instructor_company, url=instructor_web)
                        if p_created:
                            newProvider.save()
                        course.provider = newProvider
                        # AddCourseTags(course)
                        course.extractor_version = OFFLINE_EXTRACTOR_VERSION
                        course.save()

                        liveTraining = LiveTraining(course=course, max_students=max_students, min_students=min_students, session_count=session_count, prerequisites=prerequisites)
                        liveTraining.save()
            return redirect('learn:dashboard')
        else:  # form invalid
            return self.render_to_response(self.get_context_data(uploadform=OfflineContentloadForm(), formset=offc_formset, error=True))

    def form_invalid(self, mcqcreate_form, mcqa_formset):
        return self.render_to_response(self.get_context_data(form=mcqcreate_form, formset=mcqa_formset))


class OfflineCourseUpload(LoginRequiredMixin, FormView):
    raise_exception = True
    # template_name = "offline_import.html"

    def post(self, request, *args, **kwargs):
        self.object = None
        prefix = "offc"
        course_data = self.handle_uploaded_file(request.FILES.get('course_file'), request.POST, prefix)
        offlinecontentform = OfflineContentloadForm(request.POST, request.FILES)
        CourseFormSet = formset_factory(BulkCourseForm, extra=len(course_data), validate_max=True)
        offc_formset = CourseFormSet(course_data, prefix=prefix)
        template_file_name = 'offline_course_component.html'
        t = loader.get_template(template_file_name)
        context = self.get_context_data(form=offlinecontentform, formset=offc_formset)
        code = t.render(context)
        if "error_message" in request.POST:
            return HttpResponse(
                json.dumps({'code': code, 'success': False, 'error_message': request.POST["error_message"]}), content_type='application/json')
        else:
            return HttpResponse(json.dumps({'code': code, 'success': True}), content_type='application/json')
            # return self.render_to_response(context)

    def verify_headers(self, sheet):
        headers = ["ID", "Title", "Description", "Url", "Level (Beginner/Intermidiate/Advanced)", "Price", "Currency", "Max student",
                   "Min student", "Is asssessment required", "Tags", "Prerequisites", "Instructor.Title", "Instructor.Name",
                   "Instructor.LinkedIn", "Instructor.website", "Instructor.Company", "Is Available on trainer location",
                   "Primary cities", "Duration", "Is recurring", "Recurring In", "Number of times" ]
        title_list = []
        for row in sheet.iter_rows(min_row=2, max_row=2, min_col=1):
            for cell in row:
                title_list.append(cell.value)
        return all(i == j for (i, j) in zip(headers, title_list))

    def handle_uploaded_file(self, file, post, prefix):
        new_post = post.copy()
        try:
            wb = load_workbook(file)
            sheet = wb['Courses']
            if self.verify_headers(sheet):
                rows_iter = sheet.iter_rows(min_col=1, min_row=3, max_col=21, max_row=sheet.max_row)
                # vals = [[cell.value for cell in row] for row in rows_iter]
                form_index = 0
                for row in rows_iter:
                    item_prefix = prefix + "-" + str(form_index) + "-"
                    for cell in row:
                        if cell.col_idx == 2:
                            new_post[item_prefix + "title"] = str(cell.value)
                        elif cell.col_idx == 3:
                            new_post[item_prefix + "description"] = "" if cell.value==None else cell.value
                        elif cell.col_idx == 4:
                            new_post[item_prefix + "url"] = cell.value
                        elif cell.col_idx == 5:
                            new_post[item_prefix + "difficulty"] = DifficultyChoice[cell.value].value
                        elif cell.col_idx == 6:
                            new_post[item_prefix + "price_0"] = cell.value
                        elif cell.col_idx == 7:
                            new_post[item_prefix + "price_1"] = cell.value
                        elif cell.col_idx == 8:
                            new_post[item_prefix + "max_students"] = cell.value
                        elif cell.col_idx == 9:
                            new_post[item_prefix + "min_students"] = cell.value
                        elif cell.col_idx == 12:
                            new_post[item_prefix + "prerequisites"] = "" if cell.value==None else cell.value
                        elif cell.col_idx == 14:
                            new_post[item_prefix + "instructor_name"] = cell.value
                        elif cell.col_idx == 16:
                            new_post[item_prefix + "instructor_web"] = cell.value
                        elif cell.col_idx == 17:
                            new_post[item_prefix + "instructor_company"] = cell.value
                        elif cell.col_idx == 20:
                            new_post[item_prefix + "duration"] = '{:02}:00:00'.format(int(cell.value))
                        elif cell.col_idx == 23:
                            new_post[item_prefix + "session_count"] = 0 if cell.value==None else cell.value
                    form_index += 1
                new_post[prefix + '-TOTAL_FORMS'] = form_index
                new_post[prefix + '-INITIAL_FORMS'] = form_index
            else:
                new_post["error_message"] = "This sheet does not have the correct set of columns"
        except Exception as e:
            new_post["error_message"] = "Format not supported"  # e.message
            logging.getLogger('purpleskills').error(msg=file.name + ": " + e.message)

        return new_post